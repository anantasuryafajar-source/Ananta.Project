"""Migrasi HISTORI (backfill) dari sheet SKU TABEL.

Membuat ~2 tahun laba/rugi berisi data asli TANPA menimbulkan stok minus, dengan
cara memposting JURNAL RINGKASAN per bulan (bukan ribuan faktur satu-satu):

    Dr  Piutang Usaha (1-1200)      omzet
        Cr  Pendapatan Penjualan (4-1000)   omzet
    Dr  HPP (5-1000)                hpp
        Cr  Persediaan (1-1400)             hpp

Omzet = Σ(qty × harga_jual), HPP = Σ(qty × modal), dari produk yang cocok di
products.json. Ini langsung mengisi Laba Rugi, Rekap Kuartal, dan Neraca.

Alur:
  1) python -m app.import_history extract /path/ASF_MASTER_DATA.xlsx
     -> menulis seed_data/history.json (agregat bulanan)
  2) await post_history(db, company_id)  (via runner seed_history)
     -> memposting jurnal historis (idempotent: dilewati bila sudah ada)
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

SEED_DIR = Path(__file__).parent / "seed_data"

MONTHS = {
    "january": 1, "februari": 2, "february": 2, "januari": 1, "march": 3, "maret": 3,
    "april": 4, "may": 5, "mei": 5, "june": 6, "juni": 6, "july": 7, "juli": 7,
    "august": 8, "agustus": 8, "september": 9, "october": 10, "oktober": 10,
    "november": 11, "december": 12, "desember": 12,
}


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s else ""


def _price_map() -> dict[str, dict]:
    prods = json.loads((SEED_DIR / "products.json").read_text("utf-8"))
    return {_norm(p["name"]): {"sale": float(p["sale_price"]),
                               "buy": float(p["purchase_price"])} for p in prods}


def _match(name_norm: str, pmap: dict[str, dict]):
    if name_norm in pmap:
        return pmap[name_norm]
    for k, v in pmap.items():  # cocokkan awalan (nama transaksi sering dipangkas)
        if k.startswith(name_norm[:10]) or name_norm.startswith(k[:10]):
            return v
    return None


def extract_history(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "SKU TABEL" not in wb.sheetnames:
        return {"months": 0, "matched": 0, "unmatched": 0}
    pmap = _price_map()

    agg: dict[str, dict] = {}
    matched = unmatched = 0
    for row in wb["SKU TABEL"].iter_rows(min_row=2, max_col=6, values_only=True):
        cust, tgl, tahun, bulan, produk, qty = (list(row) + [None] * 6)[:6]
        if not produk or not qty:
            continue
        # tentukan tahun & bulan
        yr = None
        mo = None
        if hasattr(tgl, "year"):
            yr, mo = tgl.year, tgl.month
        else:
            try:
                yr = int(float(tahun))
            except (TypeError, ValueError):
                yr = None
            mo = MONTHS.get(str(bulan).strip().lower()) if bulan else None
        if not yr or not mo:
            continue

        price = _match(_norm(produk), pmap)
        if price is None:
            unmatched += 1
            continue
        matched += 1
        try:
            q = float(qty)
        except (TypeError, ValueError):
            continue
        key = f"{yr}-{mo:02d}"
        a = agg.setdefault(key, {"omzet": 0.0, "hpp": 0.0, "qty": 0.0})
        a["omzet"] += q * price["sale"]
        a["hpp"] += q * price["buy"]
        a["qty"] += q

    months = [{"period": k, "omzet": round(v["omzet"], 2),
               "hpp": round(v["hpp"], 2), "qty": v["qty"]}
              for k, v in sorted(agg.items())]
    (SEED_DIR / "history.json").write_text(
        json.dumps(months, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"months": len(months), "matched": matched, "unmatched": unmatched}


async def post_history(db, company_id: str) -> dict:
    """Post jurnal ringkasan bulanan dari history.json. Idempotent."""
    from datetime import date
    from decimal import Decimal
    from sqlalchemy import select
    from .models import Journal
    from .services.journal import Line, post_journal
    from .services.accounts_map import code_to_id

    path = SEED_DIR / "history.json"
    if not path.exists():
        return {"posted": 0, "skipped": 0, "note": "history.json belum ada (jalankan extract)."}
    months = json.loads(path.read_text("utf-8"))
    acc = await code_to_id(db, company_id)
    posted = skipped = 0

    for m in months:
        yr, mo = map(int, m["period"].split("-"))
        on_date = date(yr, mo, 28)
        number = f"HIST/{yr}{mo:02d}"
        exists = (await db.execute(
            select(Journal).where(Journal.company_id == company_id,
                                  Journal.number == number)
        )).scalar_one_or_none()
        if exists:
            skipped += 1
            continue
        omzet = Decimal(str(m["omzet"]))
        hpp = Decimal(str(m["hpp"]))
        if omzet <= 0:
            continue
        lines = [
            Line(acc["ar"], debit=omzet, description="Piutang historis"),
            Line(acc["sales"], credit=omzet, description="Omzet historis"),
            Line(acc["cogs"], debit=hpp, description="HPP historis"),
            Line(acc["inventory"], credit=hpp, description="Persediaan keluar historis"),
        ]
        await post_journal(
            db, company_id=company_id, number=number, on_date=on_date,
            lines=lines, memo=f"Backfill {m['period']}",
            source_type="historical",
        )
        posted += 1
    await db.commit()
    return {"posted": posted, "skipped": skipped}


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "extract":
        print("Ekstrak histori:", extract_history(sys.argv[2]))
    else:
        print(__doc__)
