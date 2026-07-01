"""Migrasi data asli ASF dari ASF_MASTER_DATA.xlsx -> data siap-seed.

Dua mode:
1)  python -m app.import_master_data extract <path.xlsx> [out_dir]
    Membaca workbook, menghasilkan JSON bersih di app/seed_data/:
      - products.json   (kompatibel ProductIn: sku,name,kind,unit,sale_price,purchase_price,min_stock)
      - contacts.json   (kompatibel ContactIn: type,name,...)
      - ongkir.json     (ekstrak biaya kurir per faktur, untuk modul kurir)
      - cashflow.json   (ekstrak buku kas, untuk laporan arus kas)
    Tidak menyentuh database — aman dijalankan di mana saja.

2)  await seed_from_master(db, company_id)
    Meng-upsert products & contacts dari products.json/contacts.json ke DB
    (dipanggil dari seed terpisah setelah CoA & company dibuat).

Catatan: sheet sumber ditulis manual sehingga berantakan (header ganda, nama
SKU tidak konsisten). Importer bersikap defensif: baris yang tidak dikenali
dilewati, bukan menggagalkan seluruh proses.
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

SEED_DIR = Path(__file__).parent / "seed_data"


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s else ""


def _num(v):
    """Angka -> float bersih, atau None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else None
    except ValueError:
        return None


# ---------------------------------------------------------------- PRODUCTS
def extract_products(wb) -> list[dict]:
    """Master produk dari sheet KOMISI (kode, nama, modal, harga jual),
    harga jual final diprioritaskan dari sheet PL (Pricelist Cash)."""
    products: dict[str, dict] = {}  # by sku code

    if "KOMISI" in wb.sheetnames:
        for row in wb["KOMISI"].iter_rows(min_row=1, max_col=8, values_only=True):
            code, _qty, name, modal, _tm, hj = (
                row[1], row[2], row[3], row[4], row[5], row[6]
            )
            if not code or not name:
                continue
            code = str(code).strip()
            name = str(name).strip()
            # lewati header & baris sampah (nama SELURUHNYA angka)
            if code.upper() in ("KODE", "") or name.upper() in ("SKU", ""):
                continue
            if re.fullmatch(r"[\d.,\s]+", name):
                continue
            if _num(modal) is None:  # baris produk asli selalu punya modal
                continue
            if code in products:  # pakai kemunculan pertama (data terlengkap)
                continue
            products[code] = {
                "sku": code,
                "name": name,
                "kind": "good",
                "unit": "botol",
                "purchase_price": _num(modal) or 0,
                "sale_price": _num(hj) or 0,
                "min_stock": 0,
            }

    # Pricelist Cash (PL) -> override sale_price bila cocok nama
    pl_by_name: dict[str, float] = {}
    if "PL" in wb.sheetnames:
        for row in wb["PL"].iter_rows(min_row=1, max_col=4, values_only=True):
            name, price = row[1], row[2]
            if name and _num(price):
                pl_by_name[_norm(name)] = _num(price)
    for p in products.values():
        hit = pl_by_name.get(_norm(p["name"]))
        if hit:
            p["sale_price"] = hit

    return list(products.values())


# ---------------------------------------------------------------- CONTACTS
def extract_contacts(wb) -> list[dict]:
    """Pelanggan unik dari kolom NAMA CUSTOMER di sheet SKU TABEL."""
    names: dict[str, str] = {}  # norm -> display
    if "SKU TABEL" in wb.sheetnames:
        for row in wb["SKU TABEL"].iter_rows(min_row=2, max_col=1, values_only=True):
            raw = row[0]
            if not raw:
                continue
            name = str(raw).strip()
            key = _norm(name)
            if not key or name.upper() == "NAMA CUSTOMER":
                continue
            names.setdefault(key, name)
    return [
        {
            "type": "customer",
            "name": n,
            "npwp": None,
            "email": None,
            "phone": None,
            "address": None,
            "payment_term_days": 30,
            "credit_limit": 0,
        }
        for n in sorted(names.values())
    ]


# ---------------------------------------------------------------- ONGKIR
def extract_ongkir(wb) -> list[dict]:
    out = []
    if "ONGKIR APRIL&MEI" not in wb.sheetnames:
        return out
    for row in wb["ONGKIR APRIL&MEI"].iter_rows(min_row=1, max_col=12, values_only=True):
        faktur = row[1] if len(row) > 1 else None
        if not faktur or not str(faktur).startswith("SI"):
            continue
        nums = [c for c in row[4:] if _num(c) is not None]
        total = _num(row[5]) if len(row) > 5 else None
        ongkir = _num(nums[-1]) if nums else None
        out.append({
            "invoice_number": str(faktur).strip(),
            "date": str(row[2])[:10] if len(row) > 2 and row[2] else None,
            "customer": str(row[4]).strip() if len(row) > 4 and row[4] else None,
            "invoice_total": total,
            "ongkir": ongkir,
        })
    return out


# ---------------------------------------------------------------- CASHFLOW
def extract_cashflow(wb) -> list[dict]:
    out = []
    if "CASHFLOW" not in wb.sheetnames:
        return out
    for row in wb["CASHFLOW"].iter_rows(min_row=3, max_col=7, values_only=True):
        bulan, kuartal, akun, tipe, nominal, kas, note = (list(row) + [None] * 7)[:7]
        if not akun or _num(nominal) is None:
            continue
        out.append({
            "month": str(bulan).strip() if bulan else None,
            "quarter": str(kuartal).strip() if kuartal else None,
            "account": str(akun).strip(),
            "account_type": str(tipe).strip() if tipe else None,
            "amount": _num(nominal),
            "cash_bucket": str(kas).strip() if kas else None,
            "note": str(note).strip() if note else None,
        })
    return out


# ---------------------------------------------------------------- RUN
def run_extract(xlsx_path: str, out_dir: str | None = None) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    data = {
        "products": extract_products(wb),
        "contacts": extract_contacts(wb),
        "ongkir": extract_ongkir(wb),
        "cashflow": extract_cashflow(wb),
    }
    out = Path(out_dir) if out_dir else SEED_DIR
    out.mkdir(parents=True, exist_ok=True)
    for key, rows in data.items():
        (out / f"{key}.json").write_text(
            json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    return {k: len(v) for k, v in data.items()}


# ---------------------------------------------------------------- SEEDER
async def seed_from_master(db, company_id: str) -> dict:
    """Upsert products & contacts dari seed_data/*.json ke DB.
    Idempotent: cocokkan produk by sku, kontak by nama (case-insensitive)."""
    from decimal import Decimal
    from sqlalchemy import select, func
    from .models import Product, Contact

    prods = json.loads((SEED_DIR / "products.json").read_text("utf-8"))
    conts = json.loads((SEED_DIR / "contacts.json").read_text("utf-8"))
    added = {"products": 0, "contacts": 0}

    for p in prods:
        exists = (await db.execute(
            select(Product).where(Product.company_id == company_id,
                                  Product.sku == p["sku"])
        )).scalar_one_or_none()
        if exists:
            continue
        db.add(Product(
            company_id=company_id, sku=p["sku"], name=p["name"],
            kind=p.get("kind", "good"), unit=p.get("unit", "botol"),
            sale_price=Decimal(str(p.get("sale_price", 0))),
            purchase_price=Decimal(str(p.get("purchase_price", 0))),
            min_stock=Decimal(str(p.get("min_stock", 0))),
        ))
        added["products"] += 1

    for c in conts:
        exists = (await db.execute(
            select(Contact).where(
                Contact.company_id == company_id,
                func.lower(Contact.name) == c["name"].lower(),
            )
        )).scalar_one_or_none()
        if exists:
            continue
        db.add(Contact(
            company_id=company_id, type=c.get("type", "customer"),
            name=c["name"], payment_term_days=c.get("payment_term_days", 30),
            credit_limit=Decimal(str(c.get("credit_limit", 0))),
        ))
        added["contacts"] += 1

    await db.commit()
    return added


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "extract":
        summary = run_extract(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
        print("Ekstrak selesai:", summary)
    else:
        print(__doc__)
